# app.py — Tableau de Bord Suivi Fabrication (TOR, synchronisation, filtres, couleurs)
# Exécuter : streamlit run App.py
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from datetime import datetime
import os
from openpyxl import load_workbook

try:
    from streamlit_autorefresh import st_autorefresh
except ImportError:
    st_autorefresh = None  # fallback si non installé

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

# -------------------------------------------------
# 0) Configuration & thème
# -------------------------------------------------
st.set_page_config(page_title="Suivi Fabrication", layout="wide")
st.title("📊 Tableau de Bord — Suivi Fabrication Structure Métallique")

# Couleurs par Étape (pour cohérence visuelle)
STEP_COLORS = {
    "Préparation": "#1f77b4",  # bleu
    "Assemblage": "#ff7f0e",  # orange
    "Traitement de surface": "#2ca02c",  # vert
    "Finalisation": "#d62728",  # rouge
    "None": "#7f7f7f"
}

# --- Thème NeoMeca (personnalisable) ---
NEOMECA_PALETTE = {
    "primary":   "#0D47A1",  # bleu foncé
    "accent":    "#FF6F00",  # orange
    "success":   "#2E7D32",  # vert
    "warning":   "#FFB300",  # jaune
    "danger":    "#D32F2F",  # rouge
    "neutral":   "#37474F",  # gris
}

# (Option) Amélioration perf légère : catégories
DTYPE_CONFIG = {
    "PHASE": "category",
    "Etape": "category",
    "PROFILE": "category",
}

# Étapes et pondérations (pour RowProgress% global)
STEPS_ORDER = ["Préparation", "Assemblage", "Traitement de surface", "Finalisation"]
STEP_RANK = {s: i for i, s in enumerate(STEPS_ORDER)}  # ordre pour logique TOR
PROGRESS_MAP = {
    "Préparation": 0.25,
    "Assemblage": 0.60,
    "Traitement de surface": 0.85,
    "Finalisation": 1.00,
    "None": 0.00
}

# -----------------------------
# Google Drive: service & utils
# -----------------------------
@st.cache_resource
def get_drive_service():
    """Construit le client Drive à partir des secrets Streamlit (service account)."""
    try:
        sa_info = dict(st.secrets["gdrive_service"])  # section TOML -> dict
    except Exception:
        return None  # pas de secrets -> pas de Drive

    scopes = ["https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
    return build("drive", "v3", credentials=creds)

def drive_get_meta(service, folder_id: str) -> dict:
    """
    Retourne les métadonnées du dossier (id, name, driveId...) pour savoir s'il
    appartient à un Drive partagé. Nécessaire pour requêtes avec corpora/driveId.
    """
    return service.files().get(
        fileId=folder_id,
        fields="id,name,driveId,parents",
        supportsAllDrives=True
    ).execute()

def drive_find_file(service, folder_id: str, name: str):
    drive_id = None
    try:
        meta = drive_get_meta(service, folder_id)
        drive_id = meta.get("driveId")  # présent si le dossier est dans un Drive partagé
    except Exception:
        pass  # fallback My Drive

    q = f"name = '{name}' and '{folder_id}' in parents and trashed = false"
    params = dict(
        q=q,
        spaces="drive",
        fields="files(id,name)",
        pageSize=1,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
    )
    if drive_id:
        params.update({"corpora": "drive", "driveId": drive_id})
    else:
        params.update({"corpora": "user"})

    res = service.files().list(**params).execute()
    files = res.get("files", [])
    return files[0] if files else None

def drive_download_excel(service, file_id: str) -> bytes:
    """Télécharge le contenu du fichier Drive (binaire) par file_id."""
    buf = BytesIO()
    req = service.files().get_media(fileId=file_id)
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    buf.seek(0)
    return buf.read()

def drive_upload_excel(service, folder_id: str, name: str, binary_data: bytes, file_id: str | None = None) -> str:
    """Met à jour (files.update) un Excel existant sur Drive. Ne crée PAS de nouveau fichier. Retourne le fileId."""
    media = MediaIoBaseUpload(
        BytesIO(binary_data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False
    )
    if file_id:
        file = service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True  # requis pour les Drives partagés
        ).execute()
        return file["id"]
    else:
        # Création interdite (conformément à la demande)
        raise RuntimeError("La création de nouveaux fichiers est désactivée.")

# --- PATCH: helper central pour mettre à jour le fichier existant

from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def update_excel_with_df(service, folder_id: str, ref_name: str,
                         df: pd.DataFrame, add_timestamp_sheet: bool = False) -> str:
    """
    Met à jour le fichier Excel existant sur Drive :
    - Réécrit/Crée la feuille 'Données' avec df
    - Optionnel : ajoute une feuille horodatée 'Sauvegarde_YYYYMMDD_HHMMSS'
    Retourne fileId.
    """
    # 1) Retrouver le fichier
    found = drive_find_file(service, folder_id, ref_name)
    if not found:
        raise FileNotFoundError(f"Fichier de référence introuvable: {ref_name}")
    file_id = found["id"]

    # 2) Charger le classeur actuel depuis Drive
    raw = drive_download_excel(service, file_id)
    wb = load_workbook(BytesIO(raw))

    # 3) Gestion de la feuille 'Données'
    has_donnees = ("Données" in wb.sheetnames)
    only_one_sheet = (len(wb.sheetnames) == 1)

    # Si 'Données' existe et qu'il y a au moins une autre feuille → on la supprime (pour réécrire proprement)
    if has_donnees and not only_one_sheet:
        ws_old = wb["Données"]
        wb.remove(ws_old)

    # Créer/obtenir la feuille 'Données'
    if "Données" in wb.sheetnames:
        ws = wb["Données"]
        # Nettoyage intégral si c’était la seule feuille
        if only_one_sheet and ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Données")

    # 4) Écrire le DataFrame dans 'Données'
    # (en-têtes + lignes; pas d’index)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # 5) Optionnel : ajouter une feuille horodatée
    if add_timestamp_sheet:
        ts_name = f"Sauvegarde_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        while ts_name in wb.sheetnames:
            ts_name += "_1"
        ts_index = len(wb.sheetnames
        ws_ts = wb.create_sheet(ts_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_ts.append(r)

    # S’assurer qu’au moins une feuille est visible & active
    ws.sheet_state = "visible"
    try:
        wb.active = wb.sheetnames.index("Données")
    except Exception:
        wb.active = 0  # fallback

    # 6) Sauvegarde : sérialiser le classeur et mettre à jour le fichier sur Drive
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    new_id = drive_upload_excel(
        service=service,
        folder_id=folder_id,
        name=ref_name,
        binary_data=buffer.read(),
        file_id=file_id  # update in-place: pas de création
    )
    return new_id

# -------------------------------
# Helper: récupérer le secret admin
# -------------------------------
def _get_admin_secret() -> str:
    # 1) variable d'environnement (simple en local)
    env_val = os.getenv("ADMIN_PASSWORD")
    if env_val:
        return env_val
    # 2) secrets (Cloud ou secrets.toml)
    try:
        return str(st.secrets["ADMIN_PASSWORD"])
    except Exception:
        return ""  # aucun secret → mode public par défaut


DEBUG_PASSWORD_HINTS = os.getenv("DEBUG_PASSWORD_HINTS", "false").lower() == "true"

# -------------------------------
# État et bandeau d'accès
# -------------------------------
st.sidebar.header("🔒 Accès")
if "is_admin" not in st.session_state:
    st.session_state["is_admin"] = False

_admin_secret = _get_admin_secret()
st.sidebar.caption(f"Mode actuel : {'Admin ✅' if st.session_state['is_admin'] else 'Public'}")

# -------------------------------
# Rendu conditionnel :
# - mode public → formulaire (avec bouton de validation)
# - mode admin → bouton de déconnexion seulement
# -------------------------------
if not st.session_state["is_admin"]:
    if not _admin_secret:
        st.sidebar.warning(
            "Aucun mot de passe admin n’est configuré.\n"
            "→ Local : ADMIN_PASSWORD (env var) ou .streamlit/secrets.toml\n"
            "→ Cloud : Settings → Secrets → ADMIN_PASSWORD"
        )

    with st.sidebar.form(key="admin_form", clear_on_submit=False):
        pwd_try = st.text_input(
            "Mot de passe admin",
            type="password",
            help="Astuce : vérifie majuscules/minuscules et les espaces.",
            placeholder="Saisir le mot de passe"
        )
        activate = st.form_submit_button("Activer le mode admin")

    if activate:
        entered = (pwd_try or "").strip()
        expected = _admin_secret.strip()
        if expected and entered == expected:
            st.session_state["is_admin"] = True
            st.sidebar.success("Mode admin activé ✅")
            st.rerun()  # ✅ API stable
        else:
            st.session_state["is_admin"] = False
            if not _admin_secret:
                st.sidebar.error("Mot de passe incorrect ❌ — aucun secret ADMIN_PASSWORD n’est configuré.")
            else:
                st.sidebar.error("Mot de passe incorrect ❌ — vérifie la casse et les espaces.")
                if DEBUG_PASSWORD_HINTS:
                    st.sidebar.info(f"Indice (non sensible) : longueur saisie = {len(entered)} caractères.")
else:
    # Mode admin : on n'affiche PAS le champ mot de passe,
    # seulement un bouton de déconnexion
    if st.sidebar.button("Se déconnecter"):
        st.session_state["is_admin"] = False
        st.sidebar.info("Mode public activé")
        st.rerun()  # ✅ API stable

# Flag unique pour la suite
is_admin = st.session_state["is_admin"]


# ---------- AUTO-SAUVEGARDE (15 min) ----------
# Timestamp d'exécution (stocké en session)
if "autosave_last_ts" not in st.session_state:
    st.session_state["autosave_last_ts"] = datetime.now().timestamp()

def _maybe_autosave():
    """Sauvegarde silencieuse toutes les 15 minutes en mode admin (réécrit 'Données' uniquement)."""
    # ✅ On lit l'état depuis la session, pas de variable globale
    is_admin_local = bool(st.session_state.get("is_admin", False))
    if not is_admin_local:
        return

    service = get_drive_service()
    folder_id = st.secrets.get("GDRIVE_FOLDER_ID", None)
    ref_name  = st.secrets.get("GDRIVE_FILE_NAME", "Structural_data.xlsx")
    if not service or not folder_id:
        return  # Drive non configuré

    now = datetime.now().timestamp()
    # 15 minutes = 900s
    if now - st.session_state["autosave_last_ts"] >= 900:
        try:
            update_excel_with_df(service, folder_id, ref_name, st.session_state["df"], add_timestamp_sheet=False)
            st.session_state["autosave_last_ts"] = now
            st.sidebar.success("💾 Auto-sauvegarde exécutée (15 min)")
        except Exception as e:
            st.sidebar.error(f"Auto-sauvegarde: {e}")

# Déclencheur front (si le composant est installé)
if st_autorefresh is not None:
    st_autorefresh(interval=15 * 60 * 1000, key="auto_refresh_5min")  # ~15 min
    _maybe_autosave()
else:
    st.sidebar.info("Installe 'streamlit-autorefresh' (pip install streamlit-autorefresh) pour l’auto-sauvegarde.")
    # Fallback manuel (optionnel)
    if bool(st.session_state.get("is_admin", False)) and st.sidebar.button("💾 Sauvegarde silencieuse (fallback)"):
        _maybe_autosave()

# -------------------------------------------------
# 1) Chargement des données
# -------------------------------------------------
DEFAULT_XLSX = "Structural_data.xlsx"  # même dossier que l'application


# NOTE: Les fonctions @st.cache_data ne doivent pas muter le df d'entrée in-place.
# Toujours retourner un nouveau DataFrame (copy / assign) pour éviter les incohérences de cache.

@st.cache_data(show_spinner=False)
def load_excel(path_or_buffer):
    # lit la première feuille automatiquement (évite les erreurs de nom)
    return pd.read_excel(path_or_buffer, engine="openpyxl")


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Colonnes attendues dans le fichier
    required = ["PHASE", "ASSEMBLY NO.", "PART NO.", "TOT MASS (Kg)", "PROFILE"]
    for c in required:
        if c not in df.columns:
            st.error(f"⚠️ Colonne manquante dans Excel : '{c}'")
            st.stop()
    # Harmonisation
    df["PHASE"] = df["PHASE"].astype(str)
    df["TOT MASS (Kg)"] = pd.to_numeric(df["TOT MASS (Kg)"], errors="coerce").fillna(0.0)
    # Colonnes d'application (si absentes)
    if "Etape" not in df.columns:
        df["Etape"] = "None"
    if "RowProgress%" not in df.columns:
        df["RowProgress%"] = 0.0
    if "CompletedMass_Row" not in df.columns:
        df["CompletedMass_Row"] = 0.0
    if "PROFILE" not in df.columns:
        df["PROFILE"] = "Standard"
    return df


# Charger fichier du dossier ou via upload
st.sidebar.header("🛠 Données")

if is_admin:
    uploaded = st.sidebar.file_uploader(
        "Importer un Excel (.xlsx)", type=["xlsx"],
        help="Optionnel : sinon 'Structural_data.xlsx' sera utilisé."
    )
else:
    uploaded = None  # pas d'upload en mode public

# --- Lecture prioritaire depuis Drive ; upload admin possible ; sinon fallback local ---
service = get_drive_service()
folder_id = st.secrets.get("GDRIVE_FOLDER_ID", None)
ref_name = st.secrets.get("GDRIVE_FILE_NAME", "Structural_data.xlsx")

try:
    if is_admin and uploaded is not None:
        # L’admin a importé un fichier → on le charge pour cette session
        df_loaded = load_excel(uploaded)
        source_label = f"Fichier importé (admin) : {uploaded.name}"
        current_source_key = f"upload::{uploaded.name}"
    elif service and folder_id:
        # Lecture Drive : fichier de référence dans le dossier partagé
        found = drive_find_file(service, folder_id, ref_name)
        if found:
            raw = drive_download_excel(service, found["id"])
            df_loaded = pd.read_excel(BytesIO(raw), engine="openpyxl")
            source_label = f"Fichier Drive : {ref_name}"
            current_source_key = f"drive::{found['id']}"
        else:
            # Pas trouvé dans Drive → fallback local (utile en dev)
            df_loaded = load_excel(DEFAULT_XLSX)
            source_label = f"Fichier local (fallback) : {DEFAULT_XLSX}"
            current_source_key = f"local::{DEFAULT_XLSX}"
    else:
        # Pas de service/secrets → fallback local
        df_loaded = load_excel(DEFAULT_XLSX)
        source_label = f"Fichier local : {DEFAULT_XLSX}"
        current_source_key = f"local::{DEFAULT_XLSX}"
except Exception as e:
    st.error(f"❌ Échec de chargement : {e}")
    st.stop()

df_loaded = ensure_columns(df_loaded)
st.caption(f"✅ {source_label}")

# -------------------------------------------------
# 1.b) Initialisation état partagé (SESSION)
# -------------------------------------------------
if "df" not in st.session_state:
    st.session_state["df"] = df_loaded.copy()
    st.session_state["source_key"] = current_source_key
elif st.session_state.get("source_key") != current_source_key:
    st.session_state["df"] = df_loaded.copy()
    st.session_state["source_key"] = current_source_key

if "refresh_needed" not in st.session_state:
    st.session_state["refresh_needed"] = False

if "dirty" not in st.session_state:
    st.session_state["dirty"] = False

# --- ÉTAT DES FILTRES ACTIFS (utilisés par filter_view) ---
if "filters" not in st.session_state:
    df_all = st.session_state["df"]
    st.session_state["filters"] = {
        "phase":   sorted(df_all["PHASE"].unique()),
        "step":    STEPS_ORDER + ["None"],
        "profile": sorted(df_all["PROFILE"].fillna("Non défini").unique()),
        "search":  "",
        "mass":    (
            float(df_all["TOT MASS (Kg)"].min()),
            float(df_all["TOT MASS (Kg)"].max())
        ),
        "sort":    "PHASE ↑",
    }

# -------------------------------------------------
# 2) Fonctions utilitaires (TOR & calculs)
# -------------------------------------------------
def recompute_progress(df: pd.DataFrame) -> pd.DataFrame:
    """Recalcule RowProgress% (pondéré par Étape) et CompletedMass_Row (masse × RowProgress%)."""
    out = df.copy()

    # 1) Masses : s'assurer d'un float
    out["TOT MASS (Kg)"] = pd.to_numeric(out["TOT MASS (Kg)"], errors="coerce").astype(float).fillna(0.0)

    # 2) Etape -> RowProgress% : caster avant map pour éviter les pièges 'category'
    #    (si 'Etape' est catégorielle, on passe par object/str ; retour en float64)
    rp = out["Etape"].astype(object).map(PROGRESS_MAP)
    out["RowProgress%"] = pd.to_numeric(rp, errors="coerce").astype(float).fillna(0.0)

    # 3) Produit vectorisé en numpy float (évite toute interférence de dtype category)
    out["CompletedMass_Row"] = (
        out["TOT MASS (Kg)"].to_numpy(dtype=float) * out["RowProgress%"].to_numpy(dtype=float)
    )

    return out

@st.cache_data(show_spinner=False)
def step_advancement(df: pd.DataFrame) -> pd.DataFrame:
    """Avancement par étape TOR — calcule la masse traitée au moins jusqu'à chaque étape."""
    total_mass = pd.to_numeric(df["TOT MASS (Kg)"], errors="coerce").fillna(0.0).sum()

    # Série de rangs numériques (int) quelle que soit la nature ddf["Etape"]
    ranks = (
        df["Etape"]
        .astype(object)                 # évite les comportements catégoriels
        .map(STEP_RANK)                 # map -> codes TOR
        .fillna(-1)
        .astype(int)                    # force dtype int pour la comparaison
    )

    rows = []
    for step in STEPS_ORDER:
        # Comparaison purement numérique (int >= int) → pas d'erreur categorical
        treated_mass = df.loc[ranks >= STEP_RANK[step], "TOT MASS (Kg)"].sum()
        pct = (treated_mass / total_mass) * 100 if total_mass > 0 else 0.0
        rows.append({"Etape": step, "CompletedMass": treated_mass, "Avancement%": pct})

    return pd.DataFrame(rows)



@st.cache_data(show_spinner=False)
def phase_advancement(df: pd.DataFrame) -> pd.DataFrame:
    """Avancement par PHASE."""
    rows = []
    for phase in sorted(df["PHASE"].unique()):
        phase_total_mass = df.loc[df["PHASE"] == phase, "TOT MASS (Kg)"].sum()
        treated_mass = df.loc[df["PHASE"] == phase, "CompletedMass_Row"].sum()
        pct = (treated_mass / phase_total_mass) * 100 if phase_total_mass > 0 else 0.0
        rows.append({"PHASE": phase, "CompletedMass": treated_mass, "Avancement%": pct})
    return pd.DataFrame(rows)

# --- Anti-régression TOR : empêcher de repasser à une étape antérieure ---
def enforce_monotonic_tor(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    # Crée une colonne technique mémorisant le rang le plus élevé déjà atteint
    if "Etape_Last" not in out.columns:
        out["Etape_Last"] = out["Etape"].map(STEP_RANK).fillna(-1)

    # Rangs courant vs dernier rang atteint
    r_now  = out["Etape"].map(STEP_RANK).fillna(-1)
    r_last = out["Etape_Last"].fillna(-1)

    # Pas de régression : on force à max(r_now, r_last)
    r_no_reg = r_now.where(r_now >= r_last, r_last)

    inv_rank = {v: k for k, v in STEP_RANK.items()}
    out["Etape"] = r_no_reg.map(inv_rank).fillna("None")
    out["Etape_Last"] = out["Etape"].map(STEP_RANK).fillna(-1)

    return out


# --- Historique en session ---
def init_history():
    if "history" not in st.session_state:
        st.session_state["history"] = []
    if "future" not in st.session_state:
        st.session_state["future"] = []

# Première recomputation
st.session_state["df"] = recompute_progress(st.session_state["df"])

# Initialisation de l’historique au chargement
init_history()

def push_history(df: pd.DataFrame):
    # On enregistre une copie légère
    st.session_state["history"].append(df.copy())
    # Une nouvelle action invalide la pile 'future'
    st.session_state["future"].clear()
    # (Option) éviter une pile trop grande :
    MAX_HIST = 10
    if len(st.session_state["history"]) > MAX_HIST:
        st.session_state["history"] = st.session_state["history"][-MAX_HIST:]

def undo():
    if st.session_state.get("history"):
        # L'état courant part dans 'future'
        st.session_state["future"].append(st.session_state["df"].copy())
        st.session_state["df"] = st.session_state["history"].pop()
        st.session_state["dirty"] = True
        try:
            step_advancement.clear()
            phase_advancement.clear()
        except Exception:
            pass
        st.success("↩️ Annulé (Undo)")

def redo():
    if st.session_state.get("future"):
        st.session_state["history"].append(st.session_state["df"].copy())
        st.session_state["df"] = st.session_state["future"].pop()
        st.session_state["dirty"] = True
        try:
            step_advancement.clear()
            phase_advancement.clear()
        except Exception:
            pass
        st.success("↪️ Rétabli (Redo)")

# -------------------------------------------------
# 3) Onglets principaux
# -------------------------------------------------
if is_admin:
    tab_edit, tab_kpi, tab_graph, tab_export = st.tabs(["✏️ Édition", "📈 KPI", "📊 Graphiques", "📤 Export"])
else:
    tab_kpi, tab_graph = st.tabs(["📈 KPI", "📊 Graphiques"])

# -------------------------------------------------
# ✏️ 3.1 Édition
# -------------------------------------------------
if is_admin:
    with tab_edit:
        st.subheader("Mise à jour des Étapes (TOR) — par pièce")

        # # --- Contrôles UX avancés (ajoute PROFILE) ---
        # phases = sorted(st.session_state["df"]["PHASE"].unique())
        # steps = STEPS_ORDER + ["None"]
        # profiles = sorted(st.session_state["df"]["PROFILE"].fillna("Non défini").unique())

        # --- FORMULAIRE DE FILTRES (validation en une fois) ---
        with st.form("filters_form", clear_on_submit=False, enter_to_submit=True):
            df_all = st.session_state["df"]
            phases = sorted(df_all["PHASE"].unique())
            steps = STEPS_ORDER + ["None"]
            profiles = sorted(df_all["PROFILE"].fillna("Non défini").unique())
            m_min = float(df_all["TOT MASS (Kg)"].min())
            m_max = float(df_all["TOT MASS (Kg)"].max())

            # Valeurs initiales = filtres actifs
            f = st.session_state["filters"]

            ph_sel = st.multiselect("Filtrer par PHASE", options=phases, default=f["phase"], key="pending_filter_phase")
            # ✅ Nouveau : case à cocher "Afficher toutes les PHASES"
            show_all_phases_default = (set(f["phase"]) == set(phases))
            show_all_phases = st.checkbox(
                "Afficher toutes les PHASES",
                value=show_all_phases_default,
                key="pending_filter_phase_all",
                help="Cochez pour ignorer la sélection ci‑dessus et afficher toutes les phases."
            )
            step_sel = st.multiselect("Filtrer par Étape", options=steps, default=f["step"], key="pending_filter_step")
            prof_sel = st.multiselect("Filtrer par PROFILE", options=profiles, default=f["profile"],
                                      key="pending_filter_profile")
            # ✅ Nouveau : case à cocher "Afficher tous les PROFILE"
            show_all_default = (set(f["profile"]) == set(profiles))
            show_all_profiles = st.checkbox(
                "Afficher tous les PROFILE",
                value=show_all_default,
                key="pending_filter_profile_all",
                help="Cochez pour ignorer la sélection ci‑dessous et afficher tous les profils."
            )

            search_q = st.text_input("🔍 Rechercher Assemblage / Pièce (contient)", value=f["search"],
                                     key="pending_filter_search")
            mass_min, mass_max = st.slider(
                "Filtrer par masse (Kg)",
                min_value=m_min, max_value=m_max,
                value=(max(m_min, f["mass"][0]), min(m_max, f["mass"][1])),
                step=0.1,
                key="pending_filter_mass",
            )
            sort_by = st.selectbox(
                "Trier par",
                options=[
                    "PHASE ↑", "PHASE ↓",
                    "Étape (rang) ↑", "Étape (rang) ↓",
                    "Masse (Kg) ↑", "Masse (Kg) ↓",
                    "PART NO. ↑", "PART NO. ↓",
                    "ASSEMBLY NO. ↑", "ASSEMBLY NO. ↓",
                ],
                index=0 if f["sort"] not in [
                    "PHASE ↑", "PHASE ↓", "Étape (rang) ↑", "Étape (rang) ↓", "Masse (Kg) ↑", "Masse (Kg) ↓",
                    "PART NO. ↑", "PART NO. ↓", "ASSEMBLY NO. ↑", "ASSEMBLY NO. ↓"
                ] else ["PHASE ↑", "PHASE ↓", "Étape (rang) ↑", "Étape (rang) ↓", "Masse (Kg) ↑", "Masse (Kg) ↓",
                        "PART NO. ↑", "PART NO. ↓", "ASSEMBLY NO. ↑", "ASSEMBLY NO. ↓"].index(f["sort"]),
                key="pending_filter_sort",
            )

            c_a, c_b = st.columns([1, 1])
            apply_filters = c_a.form_submit_button("✅ Appliquer les filtres", type="primary")
            reset_filters = c_b.form_submit_button("↺ Réinitialiser")

        # --- Application des filtres au clic (met à jour les "filtres actifs") ---
        if apply_filters:
            st.session_state["filters"] = {
                "phase":   phases  if show_all_phases   else (ph_sel   or phases),
                "step": step_sel or steps,
                "profile": profiles if show_all_profiles else (prof_sel or profiles),
                "search": (search_q or "").strip(),
                "mass": (mass_min, mass_max),
                "sort": sort_by,
            }
            st.success("✅ Filtres appliqués")

        if reset_filters:
            st.session_state["filters"] = {
                "phase": phases,
                "step": steps,
                "profile": profiles,
                "search": "",
                "mass": (m_min, m_max),
                "sort": "PHASE ↑",
            }
            st.info("↺ Filtres réinitialisés")


        # --- Filtrage vectorisé ---
        def filter_view(_df: pd.DataFrame) -> pd.DataFrame:
            f = st.session_state["filters"]

            _view = _df[
                (_df["PHASE"].isin(f["phase"])) &
                (_df["Etape"].isin(f["step"])) &
                (_df["PROFILE"].fillna("Non défini").isin(f["profile"])) &
                (_df["TOT MASS (Kg)"].between(f["mass"][0], f["mass"][1]))
                ]

            if f["search"]:
                pat = f["search"].lower()
                mask = (
                        _view["ASSEMBLY NO."].astype(str).str.lower().str.contains(pat, na=False)
                        |
                        _view["PART NO."].astype(str).str.lower().str.contains(pat, na=False)
                )
                _view = _view[mask]

            # Tri
            sort_by = f["sort"]
            if "Étape (rang)" in sort_by:
                rank = _view["Etape"].map(STEP_RANK).fillna(-1)
                asc = "↑" in sort_by
                _view = _view.assign(_rank=rank).sort_values("_rank", ascending=asc).drop(columns="_rank")
            elif "Masse (Kg)" in sort_by:
                asc = "↑" in sort_by
                _view = _view.sort_values("TOT MASS (Kg)", ascending=asc)
            elif "PHASE" in sort_by:
                asc = "↑" in sort_by
                _view = _view.sort_values("PHASE", ascending=asc)
            elif "PART NO." in sort_by:
                asc = "↑" in sort_by
                _view = _view.sort_values("PART NO.", ascending=asc)
            elif "ASSEMBLY NO." in sort_by:
                asc = "↑" in sort_by
                _view = _view.sort_values("ASSEMBLY NO.", ascending=asc)

            return _view


        st.markdown("**Éditer l’étape par pièce** (validation manuelle)")

        view_items = filter_view(st.session_state["df"])
        edit_cols  = ["PHASE", "ASSEMBLY NO.", "PART NO.", "TOT MASS (Kg)", "PROFILE", "Etape"]
        df_edit_items = view_items[edit_cols].copy()

        # ✅ 'Etape' éditable → éviter category pendant l'édition
        if "Etape" in df_edit_items.columns:
            df_edit_items["Etape"] = df_edit_items["Etape"].astype(str)

        with st.form("edit_items_form", clear_on_submit=False):
            updated_items = st.data_editor(
                df_edit_items,
                key="edit_items",
                hide_index=True,
                use_container_width=True,
                column_config={
                    "PHASE": st.column_config.TextColumn("PHASE", disabled=True, help="Phase d'appartenance"),
                    "ASSEMBLY NO.": st.column_config.TextColumn("ASSEMBLY NO.", disabled=True),
                    "PART NO.": st.column_config.TextColumn("PART NO.", disabled=True),
                    "TOT MASS (Kg)": st.column_config.NumberColumn("TOT MASS (Kg)", disabled=True, help="Masse de la pièce"),
                    "PROFILE": st.column_config.TextColumn("PROFILE", disabled=True),
                    "Etape": st.column_config.SelectboxColumn(options=STEPS_ORDER + ["None"], required=True),
                },
            )

            c1, c2 = st.columns([1, 1])
            valider = c1.form_submit_button("✅ Valider les modifications", type="primary")
            annuler = c2.form_submit_button("↩️ Annuler (réinitialiser le tableau)")

        # 3) Bouton Annuler : on ignore les edits et on relit la source
        if annuler:
            st.info("Modifications annulées.")
            st.rerun()

        # 4) Bouton Valider : on applique en UNE FOIS (merge + recompute + invalidation caches)
        if valider:
        # --- Application immédiate robuste + Undo + Anti-régression ---
            try:
                key_cols = ["ASSEMBLY NO.", "PART NO."]
                needed   = key_cols + ["Etape"]
                if all(c in updated_items.columns for c in needed):
                    updated_map = (
                        updated_items[needed]
                        .drop_duplicates()
                        .assign(Etape=lambda s: s["Etape"].fillna("None"))
                    )

                    # Historique (Undo)
                    push_history(st.session_state["df"])

                    df_base = st.session_state["df"].copy()
                    merged  = df_base.merge(updated_map, on=key_cols, how="left", suffixes=("", "_new"))
                    merged["Etape"] = merged["Etape_new"].combine_first(merged["Etape"]).fillna("None")
                    merged = merged.drop(columns=["Etape_new"])

                    # Anti-régression (si activé)
                    if st.session_state.get("anti_reg", False):
                        merged = enforce_monotonic_tor(merged)

                    # Dtypes & recompute
                    for col, dtype in DTYPE_CONFIG.items():
                        if col in merged.columns:
                            merged[col] = merged[col].astype(dtype)

                    st.session_state["df"] = recompute_progress(merged)
                    st.session_state["dirty"] = True

                    try:
                        step_advancement.clear()
                        phase_advancement.clear()
                    except Exception:
                        pass

                    st.success("✅ Modifications (pièces) appliquées")
                else:
                    st.info("ℹ️ Colonnes requises absentes (ASSEMBLY NO., PART NO., Etape) dans l’éditeur.")
            except Exception as e:
                st.error(f"❌ Erreur pendant l’application des modifications (pièces) : {e}")

# -------------------------------------------------
# 📈 3.2 KPI
# -------------------------------------------------
with tab_kpi:
    st.subheader("Indicateurs Globaux")
    total_mass = float(st.session_state["df"]["TOT MASS (Kg)"].sum())
    completed_global_mass = float(st.session_state["df"]["CompletedMass_Row"].sum())
    progress_global = (completed_global_mass / total_mass) * 100 if total_mass > 0 else 0.0
    k1, k2, k3 = st.columns(3)
    k1.metric("Masse Totale (Kg)", f"{total_mass:,.2f}")
    k2.metric("Masse Terminée (Kg)", f"{completed_global_mass:,.2f}")
    k3.metric("Avancement Global", f"{progress_global:.2f}%")

    gauge_color = (
        NEOMECA_PALETTE["success"] if progress_global >= 80
        else NEOMECA_PALETTE["warning"] if progress_global >= 50
        else NEOMECA_PALETTE["danger"]
    )
    fig_gauge = go.Figure(go.Indicator(
        mode="gauge+number",
        value=progress_global,
        title={'text': "Avancement Global (%)"},
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': gauge_color},
            'steps': [
                {'range': [0, 50], 'color': '#ffd6d6'},
                {'range': [50, 80], 'color': '#ffe9b5'},
                {'range': [80, 100], 'color': '#d6f5d6'},
            ]
        }
    ))

    st.plotly_chart(fig_gauge, use_container_width=True)
    st.divider()
    st.subheader("Avancement par Étape (TOR)")
    df_steps = step_advancement(st.session_state["df"])
    st.dataframe(
        df_steps.rename(columns={
            "Etape": "Étape",
            "CompletedMass": "Masse traitée (Kg)",
            "Avancement%": "Avancement (%)"
        }),
        use_container_width=True
    )
    fig_bar_steps = px.bar(
        df_steps,
        x="Etape", y="Avancement%",
        color="Etape",
        color_discrete_map=STEP_COLORS,
        text="Avancement%",
        title="Avancement par Étape (%)"
    )
    fig_bar_steps.update_traces(texttemplate="%{text:.2f}%", textposition="outside", marker_line_color="#333",
                                marker_line_width=0.)
    fig_bar_steps.update_yaxes(title="%", range=[0, 100])
    st.plotly_chart(fig_bar_steps, use_container_width=True)
    st.divider()
    st.subheader("Avancement par PHASE")
    df_phase = phase_advancement(st.session_state["df"])
    st.dataframe(
        df_phase.rename(columns={
            "PHASE": "Phase",
            "CompletedMass": "Masse traitée (Kg)",
            "Avancement%": "Avancement (%)"
        }),
        use_container_width=True
    )
    fig_bar_phase = px.bar(
        df_phase,
        x="PHASE", y="Avancement%",
        color="PHASE",
        text="Avancement%",
        title="Avancement par PHASE (%)"
    )
    fig_bar_phase.update_traces(texttemplate="%{text:.2f}%", textposition="outside")
    fig_bar_phase.update_yaxes(title="%", range=[0, 100])
    st.plotly_chart(fig_bar_phase, use_container_width=True)

    #Répartition masse par PHASE (camembert)
    st.subheader("Répartition de la masse par PHASE")
    df_phase_mass = st.session_state["df"].groupby("PHASE")["TOT MASS (Kg)"].sum().reset_index()
    fig_pie = px.pie(
        df_phase_mass, names="PHASE", values="TOT MASS (Kg)",
        color="PHASE", title="Masse totale par PHASE",
    )
    st.plotly_chart(fig_pie, use_container_width=True)

    #Top assemblages par masse (utile pour prioriser)
    st.subheader("Top 10 Assemblages par masse")
    df_top_asm = (st.session_state["df"]
                  .groupby("ASSEMBLY NO.")["TOT MASS (Kg)"].sum()
                  .sort_values(ascending=False).head(10).reset_index())
    fig_top_asm = px.bar(
        df_top_asm, x="ASSEMBLY NO.", y="TOT MASS (Kg)",
        title="Top 10 Assemblages (Kg)",
        color="TOT MASS (Kg)"
    )
    fig_top_asm.update_layout(xaxis_title="ASSEMBLY NO.", yaxis_title="Masse (Kg)")
    st.plotly_chart(fig_top_asm, use_container_width=True)

    #Heatmap PHASE × Étape(masse)
    st.subheader("Heatmap — Masse par PHASE × Étape")
    df_heat = (st.session_state["df"]
               .groupby(["PHASE", "Etape"])["TOT MASS (Kg)"].sum()
               .reset_index())
    pivot = df_heat.pivot(index="PHASE", columns="Etape", values="TOT MASS (Kg)").fillna(0)
    fig_heat = px.imshow(
        pivot,
        labels=dict(x="Étape", y="PHASE", color="Masse (Kg)"),
        color_continuous_scale="Blues",
        aspect="auto",
    )
    st.plotly_chart(fig_heat, use_container_width=True)

# -------------------------------------------------
# 📊 3.3 Graphiques
# -------------------------------------------------
with tab_graph:
    st.subheader("Diagramme S — Progression cumulée par Étape (TOR)")
    df_steps = step_advancement(st.session_state["df"]).copy()
    # df_steps["Avancement%"] est déjà cumulatif par étape
    df_steps["Cumul_Masse"] = df_steps["CompletedMass"]  # déjà cumulatif (par construction)

    fig_s = go.Figure()
    fig_s.add_trace(go.Scatter(
        x=df_steps["Etape"], y=df_steps["Avancement%"],
        mode="lines+markers", name="Avancement cumulé (%)",
        line=dict(width=3, color=STEP_COLORS.get("Préparation", "#1f77b4"))
    ))
    fig_s.add_trace(go.Bar(
        x=df_steps["Etape"], y=df_steps["Cumul_Masse"],
        name="Masse cumulée (Kg)", marker_color="#9ecae1", opacity=0.6, yaxis="y2"
    ))
    fig_s.update_layout(
        title="Diagramme S — % cumulé & masse cumulée",
        yaxis=dict(title="% cumulé", range=[0, 100]),
        yaxis2=dict(title="Masse (Kg)", overlaying="y", side="right"),
        legend=dict(orientation="h")
    )
    st.plotly_chart(fig_s, use_container_width=True)

# -------------------------------------------------
# 📤 3.4 Export
# -------------------------------------------------
if is_admin:
    with tab_export:
        st.subheader("Sauvegarde Drive — réécrit 'Données' et ajoute une feuille horodatée")
        service = get_drive_service()
        folder_id = st.secrets.get("GDRIVE_FOLDER_ID", None)
        ref_name  = st.secrets.get("GDRIVE_FILE_NAME", "Structural_data.xlsx")

        if not service or not folder_id:
            st.error("⚠️ Drive non configuré. Ajoute les secrets [gdrive_service] + GDRIVE_FOLDER_ID + GDRIVE_FILE_NAME.")
        else:
            # --- BOUTON UNIQUE DE SAUVEGARDE ---
            if st.button("💾 Sauvegarder sur Drive (écraser la référence) + feuille horodatée", type="primary"):
                try:
                    new_id = update_excel_with_df(
                        service=service,
                        folder_id=folder_id,
                        ref_name=ref_name,
                        df=st.session_state["df"],
                        add_timestamp_sheet=True  # crée une nouvelle feuille horodatée à chaque sauvegarde
                    )
                    st.success(f"✅ Fichier mis à jour (fileId={new_id}). 'Données' réécrite + feuille horodatée ajoutée.")
                except Exception as e:
                    st.error(f"❌ Échec de la sauvegarde Drive : {e}")

            # --- Téléchargement local (optionnel ; inchangé) ---
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as w:
                st.session_state["df"].to_excel(w, index=False, sheet_name="Données")
            buffer.seek(0)
            st.download_button(
                label="⬇️ Télécharger (Excel modifié)",
                data=buffer,
                file_name=f"Suivi_Fabrication_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
